"""
reports.writer — Excel writer for the v2 combined workbook.

Strategy: call v1's reporter.build_workbook to write the v1 sheets first
(reuses v1's exact styling and column conventions), then append the v2
sheets via openpyxl. The combined file's v1 sheets are functionally
identical to v1's standalone output, but may not be byte-identical due to
openpyxl rewriting on re-save — only v1's standalone output is
regression-tested.

New sheets appended:
  Combined_Market_View    Joined per-(CIP, AWLEVEL) view: v1 columns + labor
  Labor_By_CIP_State      Per-(CIP, state) labor aggregator output
  Labor_Flat              Per-(CIP, SOC, state) drilldown (flat mode)
  Unmatched_CIPs          CIPs with no SOC mapping in the crosswalk
  Disclosure              Vintage info + methodology footer
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# v1 reporter is reused for the v1 sheets.
from reporter import AWLEVEL_LABELS, build_workbook  # type: ignore[import-not-found]

from reports.labels import ALL_LABELS, rename_for_display


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WRAP = Alignment(wrap_text=True, vertical='top')
MAX_COL_WIDTH = 50


def _style_sheet(ws: Worksheet) -> None:
    """Match v1's basic styling: bold gray header, freeze top row, autosize."""
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COL_WIDTH)


# Columns that must be written as text to preserve formatting (leading zeros
# in CIPCODE, dashed SOC formatting). pandas otherwise reads them back as
# numeric and "01.0508" becomes 1.0508 — a different CIP code entirely.
# Includes BOTH the raw column names AND their post-rename display labels
# (since labels are applied before we write to the sheet).
_FORCE_TEXT_COLUMNS = frozenset({
    # Raw column names
    'CIPCODE', 'CIPTitle', 'SOCCode', 'SOCTitle',
    'PRIM_STATE', 'AREA_TITLE', 'state_abbr', 'state_name',
    'linked_socs', 'LABOR_LINKED_SOCS_SEL', 'LABOR_LINKED_SOCS_NAT',
    'crosswalk_status', 'aggregation_mode', 'vintage',
    'NOTE', 'area_kind', 'suppression_flag',
    # Display labels (after rename_for_display)
    'CIP Code', 'CIP Title', 'SOC Code', 'SOC Title',
    'State', 'Area', 'Geography', 'Source', 'Metric',
    'Linked SOC Codes', 'Crosswalk Status', 'Aggregation Mode',
    'Source Vintage', 'Area Type', 'Suppression Flag',
    'Linked SOCs — Selected', 'Linked SOCs — National',
    'Primary State', 'Award Level',
})


def _write_df_to_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Write a DataFrame to a new sheet in the workbook."""
    if sheet_name in wb.sheetnames:
        # Replace existing sheet by removing and recreating.
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    # Header
    if df.empty:
        ws.append(['(no rows)'])
        return
    columns = list(df.columns)
    ws.append(columns)
    # Identify columns that need text format to preserve leading zeros etc.
    text_col_idxs = {
        idx + 1 for idx, col in enumerate(columns)
        if col in _FORCE_TEXT_COLUMNS
    }
    # Data rows
    for _, row in df.iterrows():
        ws.append([
            (None if pd.isna(v) else v)
            for v in row.tolist()
        ])
    # Apply text format to the force-text columns so Excel + pandas-on-read
    # treat them as strings. number_format='@' is openpyxl's text format.
    if text_col_idxs:
        max_row = ws.max_row
        for col_idx in text_col_idxs:
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                # Force string representation. If the value is already a str
                # this is a no-op. If openpyxl inferred numeric, this re-stringifies.
                if cell.value is not None:
                    cell.value = str(cell.value)
                cell.number_format = '@'
    _style_sheet(ws)


def _write_disclosure(
    wb, vintage: dict, primary_state: str, label: str,
    saturation_caveat: Optional[str] = None,
) -> None:
    """Write the disclosure / methodology sheet — single-cell narrative + table."""
    if 'Disclosure' in wb.sheetnames:
        del wb['Disclosure']
    ws = wb.create_sheet('Disclosure')

    # Title block
    ws.append(['IPEDS + Labor Market — Disclosure & Methodology'])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    # Vintage table
    ws.append(['Field', 'Value'])
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    rows = [
        ('Report label', label),
        ('Primary geography', primary_state),
        ('Aggregation mode', vintage.get('aggregation_mode', '')),
        ('OEWS vintage', vintage.get('oews', '')),
        ('BLS Projections vintage', vintage.get('projections', '')),
        ('CA EDD vintage', vintage.get('edd', '')),
        ('Census ACS vintage', vintage.get('census', '')),
        ('CIP-SOC crosswalk', vintage.get('crosswalk', '')),
    ]
    for k, v in rows:
        ws.append([k, str(v)])

    ws.append([])
    ws.append(['Full disclosure statement (paste into reports as a footer):'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([vintage.get('disclosure', '')])
    ws.cell(row=ws.max_row, column=1).alignment = WRAP
    disclosure_row = ws.max_row

    saturation_row = None
    if saturation_caveat:
        ws.append([])
        ws.append(['Saturation ratio caveat:'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([saturation_caveat])
        ws.cell(row=ws.max_row, column=1).alignment = WRAP
        saturation_row = ws.max_row

    # Column widths
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 80
    # Wrap the long narrative rows to a few reasonable lines.
    ws.row_dimensions[disclosure_row].height = 80
    if saturation_row:
        ws.row_dimensions[saturation_row].height = 110
    ws.freeze_panes = 'A4'


def write_combined_workbook(
    dataset: dict,
    output_dir: Path,
) -> Path:
    """
    Write the combined v1 + v2 Excel workbook.

    `dataset` is what `combine.build_combined_dataset` returns.
    Returns the path of the written file.
    """
    completions = dataset['completions']
    labor = dataset['labor']
    primary_state = dataset['primary_state']
    label = dataset['label']

    # Step 1 — v1 sheets via v1's own writer. Use a label that flags combined.
    # For the Market_View sheet, feed the AUGMENTED long frame (v1's 4 rows
    # per program + 2 reported-count rows). v1's writer just renders whatever
    # rows it's given; v1's own CLI still passes its unaugmented frame, so the
    # regression-locked v1 output is unaffected.
    combined_label = f'COMBINED_{label}'
    market_view_for_excel = dataset.get(
        'market_view_long_augmented', completions['market_view_long'],
    )
    out_path = build_workbook(
        output_dir=output_dir,
        label=combined_label,
        institutions_df=completions['institutions_view'],
        completions_by_year=completions['filtered'],
        cagr_df=completions['cagr_df'],
        market_view_df=market_view_for_excel,
        varlist_df=completions['varlist_df'],
        only_latest_completions=False,
    )

    # Step 2 — append v2 sheets to the file v1 just wrote.
    wb = load_workbook(out_path)

    # 2a. Long-format labor view — the primary labor sheet. Already has
    # plain-English column names because it's produced by build_labor_long_view.
    _write_df_to_sheet(wb, 'Labor_View_Long', dataset.get('labor_long', pd.DataFrame()))

    # 2b. Per-(CIP × state) labor drilldown — filtered to user's CIPs,
    # renamed for readability.
    labor_drilldown = rename_for_display(
        dataset.get('labor_aggregated_filtered', labor['aggregated'])
    )
    # Award level on the drilldown sheet — labor data isn't AWLEVEL-keyed,
    # but a user dropping into this sheet from the Combined view may want
    # the v1 vocabulary. No transform needed — drilldown is per-CIP per-state.
    _write_df_to_sheet(wb, 'Labor_Detail_by_State', labor_drilldown)

    # 2c. Combined per-(CIP × AWLEVEL) view — keep as a "wide drilldown"
    # but with labels applied AND Award Level numeric codes mapped.
    cmv = dataset.get('combined_market_view', pd.DataFrame()).copy()
    if not cmv.empty and 'AWLEVEL' in cmv.columns:
        cmv['Award Level'] = cmv['AWLEVEL'].map(
            lambda v: AWLEVEL_LABELS.get(int(v), str(v)) if pd.notna(v) else ''
        )
        cmv = cmv.drop(columns=['AWLEVEL'])
        # Move Award Level next to CIP Title for readability
        cols = list(cmv.columns)
        if 'Award Level' in cols and 'CIPTitle' in cols:
            cols.remove('Award Level')
            insert_at = cols.index('CIPTitle') + 1
            cols.insert(insert_at, 'Award Level')
            cmv = cmv[cols]
    cmv = rename_for_display(cmv)
    _write_df_to_sheet(wb, 'Combined_Wide_Drilldown', cmv)

    # 2d. SOC-level flat drilldown — for analysts who want the per-SOC source.
    flat_renamed = rename_for_display(labor['flat'])
    _write_df_to_sheet(wb, 'Labor_Flat_SOC_Level', flat_renamed)

    # 2e. Saturation — per-CIP completions vs openings (long format,
    # already plain-English columns). Placed before Unmatched so it sits
    # near the headline labor sheets.
    _write_df_to_sheet(wb, 'Saturation_by_CIP', dataset.get('saturation', pd.DataFrame()))

    # 2f. Unmatched CIPs — filtered to user's CIPs if a filter is set.
    unmatched_list = dataset.get('unmatched_filtered', labor['unmatched_cips'])
    unmatched_df = pd.DataFrame({
        'CIP Code': unmatched_list,
        'Note': ['No SOC mapping in the NCES CIP-SOC crosswalk'] * len(unmatched_list),
    })
    _write_df_to_sheet(wb, 'Unmatched_CIPs', unmatched_df)

    _write_disclosure(
        wb, labor['vintage'], primary_state, label,
        saturation_caveat=dataset.get('saturation_caveat'),
    )

    wb.save(out_path)
    return out_path
