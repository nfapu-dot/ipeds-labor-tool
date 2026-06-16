"""
reporter.py — Builds the Excel workbook output.

Output path:
  output/reports/IPEDS_Completions_{state_or_custom}_{timestamp}.xlsx

Tab order:
  1. Institutions
  2-6. Completions_{year}        (one per year, sorted UNITID/CIPCODE/AwardLevel)
  7. CAGR_by_Institution         (1 row per UNITID×CIP×AWLEVEL, wide on year)
  8. Market_View                 (1 row per CIP×AWLEVEL, selected + national)
  9. Definitions                 (variables + AWLEVEL/CONTROL/ICLEVEL/Carnegie + flag codes)

Formatting (applied to every tab):
- Freeze top row + first two columns (openpyxl 'C2').
- Bold header row, light gray fill (#F2F2F2).
- Integer format (#,##0) on count columns.
- Percentage format (0.0%) on CAGR / % change columns.
- Auto-fit column widths (max 45 chars).

Conditional formatting:
- CAGR tab: direct fill — green > 0, red < 0, gray N/A.
- Program Growth: 3-color scale on % change column.
- Market View: 3-color scale on Market CAGR column.

NaN handling: empty cells everywhere — preserves IPEDS <3 suppression.
"""
from __future__ import annotations

import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

console = Console()

# ── styling constants ──
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
INT_FORMAT = '#,##0'
PCT_FORMAT = '0.0%'
MAX_COL_WIDTH = 45

CAGR_POS_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')  # light green
CAGR_NEG_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')  # light red
CAGR_NA_FILL = PatternFill(fill_type='solid', fgColor='D9D9D9')   # light gray

# 3-color scale: red (low) → yellow (mid=0) → green (high)
SCALE_RED = 'F8696B'
SCALE_YELLOW = 'FFEB84'
SCALE_GREEN = '63BE7B'

# ── lookup tables (for the Definitions tab + Institutions/Completions labels) ──
CONTROL_LABELS = {1: 'Public', 2: 'Private nonprofit', 3: 'Private for-profit'}
ICLEVEL_LABELS = {1: '4-year', 2: '2-year', 3: 'Less-than-2-year'}

# Spec §Data Inputs lists 1–11. IPEDS files also use 17/18/19/20 in practice
# (legacy doctorate composite + post-2010 doctorate subtypes). Documented here
# so the Definitions tab is complete.
AWLEVEL_LABELS = {
    1: 'Postsecondary award < 1 year',
    2: 'Postsecondary award 1–2 years',
    3: "Associate's degree",
    4: 'Postsecondary award 2–4 years',
    5: "Bachelor's degree",
    6: 'Post-baccalaureate certificate',
    7: "Master's degree",
    8: "Post-master's certificate",
    9: 'Doctorate — research/scholarship',
    10: 'Doctorate — professional practice',
    11: 'Doctorate — other',
    17: "Doctor's degree (legacy composite, pre-2010)",
    18: "Doctor's degree — research/scholarship (legacy code)",
    19: "Doctor's degree — professional practice (legacy code)",
    20: "Doctor's degree — other (legacy code)",
}

CAGR_FLAG_DEFINITIONS = [
    ('OK', 'Both endpoints > 0; CAGR computed normally.'),
    ('New Program', 'Start completions = 0; CAGR undefined (cannot divide).'),
    ('Program Ended', 'Start > 0 and end = 0; CAGR shown as -100%.'),
    ('Missing Data', 'Either endpoint is suppressed (<3) or absent from file; CAGR N/A.'),
]

# 2021 Carnegie Basic Classification — the C21BASIC column in HD files holds these codes.
# Source: Carnegie Classifications, American Council on Education.
CARNEGIE_LABELS = {
    -2: 'Not classified / not applicable',
    1:  "Associate's Colleges: High Transfer-High Traditional",
    2:  "Associate's Colleges: High Transfer-Mixed Traditional/Nontraditional",
    3:  "Associate's Colleges: High Transfer-High Nontraditional",
    4:  "Associate's Colleges: Mixed Transfer/Career & Technical-High Traditional",
    5:  "Associate's Colleges: Mixed Transfer/Career & Technical-Mixed",
    6:  "Associate's Colleges: Mixed Transfer/Career & Technical-High Nontraditional",
    7:  "Associate's Colleges: High Career & Technical-High Traditional",
    8:  "Associate's Colleges: High Career & Technical-Mixed",
    9:  "Associate's Colleges: High Career & Technical-High Nontraditional",
    10: 'Special Focus Two-Year: Health Professions',
    11: 'Special Focus Two-Year: Technical Professions',
    12: 'Special Focus Two-Year: Arts & Design',
    13: 'Special Focus Two-Year: Other Fields',
    14: "Baccalaureate/Associate's Colleges: Mixed Baccalaureate/Associate's",
    15: 'Doctoral Universities: Very High Research Activity (R1)',
    16: 'Doctoral Universities: High Research Activity (R2)',
    17: 'Doctoral/Professional Universities',
    18: "Master's Colleges & Universities: Larger Programs (M1)",
    19: "Master's Colleges & Universities: Medium Programs (M2)",
    20: "Master's Colleges & Universities: Smaller Programs (M3)",
    21: 'Baccalaureate Colleges: Arts & Sciences Focus',
    22: 'Baccalaureate Colleges: Diverse Fields',
    23: "Baccalaureate/Associate's Colleges: Associate's Dominant",
    24: 'Special Focus Four-Year: Faith-Related Institutions',
    25: 'Special Focus Four-Year: Medical Schools & Centers',
    26: 'Special Focus Four-Year: Other Health Professions Schools',
    27: 'Special Focus Four-Year: Research Institutions',
    28: 'Special Focus Four-Year: Engineering & Technology-Related Schools',
    29: 'Special Focus Four-Year: Other Technology-Related Schools',
    30: 'Special Focus Four-Year: Business & Management Schools',
    31: 'Special Focus Four-Year: Arts, Music & Design Schools',
    32: 'Special Focus Four-Year: Law Schools',
    33: 'Special Focus Four-Year: Other Special Focus Institutions',
}


# ---------------------------------------------------------------------------
# Top-level builder
# ---------------------------------------------------------------------------

def build_workbook(
    output_dir: Path,
    label: str,
    institutions_df: pd.DataFrame,
    completions_by_year: Dict[int, pd.DataFrame],
    cagr_df: Optional[pd.DataFrame] = None,
    program_growth_df: Optional[pd.DataFrame] = None,  # retained for back-compat; ignored
    market_view_df: Optional[pd.DataFrame] = None,
    varlist_df: Optional[pd.DataFrame] = None,
    only_latest_completions: bool = False,
) -> Path:
    """
    Build and save the Excel workbook. Returns the output path.

    The per-institution Program_Growth tab was removed (user feedback: not the
    intended summary). The CIP-level breadth/volume picture lives in Market_View,
    which now includes national-comparison columns when present.

    `program_growth_df` is accepted for back-compat with older callers but
    ignored — no tab is produced from it.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_label = ''.join(c if c.isalnum() else '_' for c in (label or '')).strip('_') or 'custom'
    out_path = output_dir / f'IPEDS_Completions_{safe_label}_{timestamp}.xlsx'

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    _build_institutions_tab(wb, institutions_df)

    years = sorted(completions_by_year.keys())
    if only_latest_completions and years:
        years = [years[-1]]
    for y in years:
        _build_completions_tab(wb, y, completions_by_year[y])

    if cagr_df is not None:
        _build_cagr_tab(wb, cagr_df)
    if market_view_df is not None:
        _build_market_view_tab(wb, market_view_df)

    _build_definitions_tab(wb, varlist_df)

    wb.save(out_path)
    console.log(f'[green]wrote workbook[/] → {out_path}')
    console.log(f'  sheets: {wb.sheetnames}')
    return out_path


# ---------------------------------------------------------------------------
# Institutions tab
# ---------------------------------------------------------------------------

def _prepare_institutions_view(institutions_df: pd.DataFrame) -> pd.DataFrame:
    df = institutions_df.copy()
    out = pd.DataFrame()
    out['UNITID'] = df.get('UNITID')
    out['Institution'] = df.get('INSTNM')
    out['State'] = df.get('STABBR')
    out['Control'] = df.get('CONTROL')
    out['Control Label'] = (
        df['CONTROL'].map(CONTROL_LABELS) if 'CONTROL' in df.columns else None
    )
    out['Level'] = df.get('ICLEVEL')
    out['Level Label'] = (
        df['ICLEVEL'].map(ICLEVEL_LABELS) if 'ICLEVEL' in df.columns else None
    )
    out['Carnegie Code'] = df.get('CARNEGIE')
    out['Carnegie Classification'] = (
        df['CARNEGIE'].map(CARNEGIE_LABELS) if 'CARNEGIE' in df.columns else None
    )
    out['HD Source Year'] = df.get('HD_SOURCE_YEAR')
    if 'UNITID' in out.columns:
        out = out.sort_values('UNITID', kind='stable')
    return out.reset_index(drop=True)


def _build_institutions_tab(wb: Workbook, institutions_df: pd.DataFrame) -> None:
    view = _prepare_institutions_view(institutions_df)
    ws = wb.create_sheet('Institutions')
    _write_dataframe(ws, view)
    _format_sheet(
        ws,
        integer_columns=['UNITID', 'Control', 'Level', 'Carnegie Code', 'HD Source Year'],
    )


# ---------------------------------------------------------------------------
# Completions_{year} tabs
# ---------------------------------------------------------------------------

def _prepare_completions_view(completions_df: pd.DataFrame) -> pd.DataFrame:
    df = completions_df.copy()
    out = pd.DataFrame()
    out['UNITID'] = df.get('UNITID')
    out['Institution'] = df.get('INSTNM')
    out['CIPCODE'] = df.get('CIPCODE')
    out['CIP Title'] = df.get('CIPTitle')
    out['Award Level'] = df.get('AWLEVEL')
    out['Total'] = df.get('CTOTALT')
    out['Men'] = df.get('CTOTALM')
    out['Women'] = df.get('CTOTALW')
    sort_cols = [c for c in ('UNITID', 'CIPCODE', 'Award Level') if c in out.columns]
    if sort_cols:
        out = out.sort_values(sort_cols, kind='stable')
    return out.reset_index(drop=True)


def _build_completions_tab(wb: Workbook, year: int, completions_df: pd.DataFrame) -> None:
    view = _prepare_completions_view(completions_df)
    ws = wb.create_sheet(f'Completions_{year}')
    _write_dataframe(ws, view)
    _format_sheet(
        ws,
        integer_columns=['UNITID', 'Award Level', 'Total', 'Men', 'Women'],
    )


# ---------------------------------------------------------------------------
# CAGR_by_Institution tab
# ---------------------------------------------------------------------------

def _awlevel_to_label(v):
    """Map AWLEVEL int to '5 — Bachelor\\'s degree'. Blank for NaN."""
    if v is None or (isinstance(v, float) and pd.isna(v)) or v is pd.NA:
        return ''
    try:
        code = int(v)
    except (TypeError, ValueError):
        return str(v)
    label = AWLEVEL_LABELS.get(code)
    return f'{code} — {label}' if label else f'{code}'


def _prepare_cagr_view(cagr_df: pd.DataFrame) -> pd.DataFrame:
    """Rename CTOTALT_{year} → "{year} Completions" for friendly headers."""
    df = cagr_df.copy()
    rename = {
        c: f'{c.split("_")[-1]} Completions'
        for c in df.columns if c.startswith('CTOTALT_')
    }
    df = df.rename(columns=rename)
    if 'AWLEVEL' in df.columns:
        df = df.rename(columns={'AWLEVEL': 'Award Level'})
    if 'INSTNM' in df.columns:
        df = df.rename(columns={'INSTNM': 'Institution'})
    if 'STABBR' in df.columns:
        df = df.rename(columns={'STABBR': 'State'})
    if 'CIPTitle' in df.columns:
        df = df.rename(columns={'CIPTitle': 'CIP Title'})
    if 'Award Level' in df.columns:
        df['Award Level'] = df['Award Level'].apply(_awlevel_to_label)
    return df


def _build_cagr_tab(wb: Workbook, cagr_df: pd.DataFrame) -> None:
    view = _prepare_cagr_view(cagr_df)
    ws = wb.create_sheet('CAGR_by_Institution')
    _write_dataframe(ws, view)

    year_cols = [c for c in view.columns if c.endswith(' Completions')]
    # Award Level is now a string ("5 — Bachelor's degree") so it's not an int col.
    _format_sheet(
        ws,
        integer_columns=['UNITID'] + year_cols,
        percent_columns=['CAGR'],
    )

    # CAGR direct-fill: green > 0, red < 0, gray for blank.
    if 'CAGR' in view.columns:
        col_idx = list(view.columns).index('CAGR') + 1
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:  # skip header
            v = cell.value
            if v is None:
                cell.fill = CAGR_NA_FILL
            elif isinstance(v, (int, float)):
                if v > 0:
                    cell.fill = CAGR_POS_FILL
                elif v < 0:
                    cell.fill = CAGR_NEG_FILL
                # v == 0: leave unfilled


# ---------------------------------------------------------------------------
# Market_View tab
# ---------------------------------------------------------------------------

def _build_market_view_tab(wb: Workbook, mv_long_df: pd.DataFrame) -> None:
    """
    Long-format Market View. Caller passes the output of
    aggregator.market_view_to_long() — 4 rows per (CIP × Award Level):
    Selected/Completions, Selected/Programs, National/Completions, National/Programs.

    Columns: CIPCODE, CIP Title, Award Level, Geography, Metric,
             {start_year}, …, {end_year}, CAGR, Flag

    "Programs" = number of institutions offering this CIP × Award Level
    (one institution offering = one program offering).
    """
    if mv_long_df is None or mv_long_df.empty:
        ws = wb.create_sheet('Market_View')
        ws.cell(row=1, column=1, value='No data matches the configured filters.')
        return

    view = mv_long_df.copy()
    if 'Award Level' in view.columns:
        view['Award Level'] = view['Award Level'].apply(_awlevel_to_label)

    ws = wb.create_sheet('Market_View')
    _write_dataframe(ws, view)

    year_cols = [c for c in view.columns if c.isdigit()]
    # Award Level is now a labeled string — not an integer column.
    _format_sheet(
        ws,
        integer_columns=year_cols,
        percent_columns=['CAGR'],
    )

    if 'CAGR' in view.columns and ws.max_row > 1:
        col_idx = list(view.columns).index('CAGR') + 1
        col_letter = get_column_letter(col_idx)
        rng = f'{col_letter}2:{col_letter}{ws.max_row}'
        ws.conditional_formatting.add(rng, _three_color_scale())


def _three_color_scale() -> ColorScaleRule:
    """Red (min) → yellow (0) → green (max)."""
    return ColorScaleRule(
        start_type='min', start_color=SCALE_RED,
        mid_type='num', mid_value=0, mid_color=SCALE_YELLOW,
        end_type='max', end_color=SCALE_GREEN,
    )


# ---------------------------------------------------------------------------
# Definitions tab
# ---------------------------------------------------------------------------

def _prepare_definitions(varlist_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []

    # Section 1 — Variables (from the loaded IPEDS varlist)
    if varlist_df is not None and not varlist_df.empty:
        for _, row in varlist_df.iterrows():
            rows.append({
                'Category': 'Variable',
                'Code': str(row.get('varName') or ''),
                'Label': str(row.get('varTitle') or ''),
                'Source': 'IPEDS varlist',
            })

    # Section 2 — Award Level codes
    for code, label in AWLEVEL_LABELS.items():
        rows.append({
            'Category': 'Award Level',
            'Code': code,
            'Label': label,
            'Source': 'SPEC §AWLEVEL + IPEDS legacy codes',
        })

    # Section 3 — CONTROL codes
    for code, label in CONTROL_LABELS.items():
        rows.append({
            'Category': 'Control',
            'Code': code,
            'Label': label,
            'Source': 'SPEC §HD',
        })

    # Section 4 — ICLEVEL codes
    for code, label in ICLEVEL_LABELS.items():
        rows.append({
            'Category': 'Level (ICLEVEL)',
            'Code': code,
            'Label': label,
            'Source': 'SPEC §HD',
        })

    # Section 5 — Carnegie codes (2021 Basic Classification)
    for code in sorted(CARNEGIE_LABELS.keys()):
        rows.append({
            'Category': 'Carnegie Classification',
            'Code': code,
            'Label': CARNEGIE_LABELS[code],
            'Source': '2021 Carnegie Basic Classification (C21BASIC)',
        })

    # Section 6 — CAGR flag values
    for code, label in CAGR_FLAG_DEFINITIONS:
        rows.append({
            'Category': 'CAGR Flag',
            'Code': code,
            'Label': label,
            'Source': 'SPEC §Calculations',
        })

    return pd.DataFrame(rows, columns=['Category', 'Code', 'Label', 'Source'])


def _build_definitions_tab(wb: Workbook, varlist_df: Optional[pd.DataFrame]) -> None:
    view = _prepare_definitions(varlist_df)
    ws = wb.create_sheet('Definitions')
    _write_dataframe(ws, view)
    _format_sheet(ws)
    # Wrap long Label values
    if 'Label' in view.columns:
        col_idx = list(view.columns).index('Label') + 1
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


# ---------------------------------------------------------------------------
# Shared write + format primitives
# ---------------------------------------------------------------------------

def _write_dataframe(ws: Worksheet, df: pd.DataFrame) -> None:
    """Write headers in row 1 and rows starting row 2. NaN → empty cell."""
    headers: List[str] = list(df.columns)
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, h in enumerate(headers, start=1):
            v = row[h]
            if v is None or v is pd.NA or (isinstance(v, float) and pd.isna(v)):
                continue
            if hasattr(v, 'item'):
                try:
                    v = v.item()
                except (ValueError, TypeError):
                    pass
            ws.cell(row=r_idx, column=c_idx, value=v)


def _format_sheet(
    ws: Worksheet,
    integer_columns: Optional[List[str]] = None,
    percent_columns: Optional[List[str]] = None,
) -> None:
    integer_columns = integer_columns or []
    percent_columns = percent_columns or []

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.freeze_panes = 'C2'

    header_idx = {cell.value: cell.column for cell in ws[1]}
    for col_name in integer_columns:
        col_idx = header_idx.get(col_name)
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:
            cell.number_format = INT_FORMAT
    for col_name in percent_columns:
        col_idx = header_idx.get(col_name)
        if col_idx is None:
            continue
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter][1:]:
            cell.number_format = PCT_FORMAT

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            s = str(cell.value)
            if len(s) > max_len:
                max_len = len(s)
        if max_len:
            ws.column_dimensions[col_letter].width = min(MAX_COL_WIDTH, max_len + 2)


# ---------------------------------------------------------------------------
# Smoke test
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    from loader import load_all, load_years_config
    from joiner import build_institution_metadata, build_cip_title_lookup, join_all_years
    from aggregator import (
        apply_filters, compute_cagr_table,
        compute_program_growth, compute_market_view,
    )

    project_root = Path(__file__).resolve().parent.parent
    years_cfg = load_years_config(project_root / 'config' / 'years.yaml')

    console.rule('[bold]reporter — smoke test (stage 2: all tabs)[/]')
    console.print('[dim]loading…[/]')
    loaded = load_all(
        years_cfg,
        raw_dir=project_root / 'data' / 'raw',
        dict_dir=project_root / 'data' / 'dictionary',
    )
    metadata = build_institution_metadata(loaded['hd'])
    cip_titles = build_cip_title_lookup(loaded['crosswalk'])

    selected_unitids = [110635, 110662, 209612, 236577, 101541]
    cip_codes = ['51.3801', '52.0201', '11.0701', '14.0901']
    award_levels = [5, 7]

    joined = join_all_years(
        loaded['ca'], metadata, cip_titles,
        selected_unitids=selected_unitids, quiet=True,
    )
    filtered = apply_filters(
        joined, cip_codes=cip_codes, award_levels=award_levels,
        include_residual=False, quiet=True,
    )
    cagr_df = compute_cagr_table(
        filtered,
        start_year=years_cfg['cagr_start_year'],
        end_year=years_cfg['cagr_end_year'],
    )
    pg_df = compute_program_growth(
        filtered, metadata,
        start_year=years_cfg['cagr_start_year'],
        end_year=years_cfg['cagr_end_year'],
    )
    mv_df = compute_market_view(
        filtered,
        start_year=years_cfg['cagr_start_year'],
        end_year=years_cfg['cagr_end_year'],
    )

    institutions_view = metadata[metadata['UNITID'].isin(selected_unitids)]

    out_path = build_workbook(
        output_dir=project_root / 'output' / 'reports',
        label='full_smoke_test',
        institutions_df=institutions_view,
        completions_by_year=filtered,
        cagr_df=cagr_df,
        program_growth_df=pg_df,
        market_view_df=mv_df,
        varlist_df=loaded['varlist'],
        only_latest_completions=False,
    )

    # Verify by re-reading
    console.rule('verify')
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n_rows = ws.max_row - 1  # minus header
        n_cols = ws.max_column
        headers = tuple(c.value for c in ws[1])
        cf_rules = sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())
        console.print(
            f'[cyan]{sheet_name:25s}[/]  {n_rows:4d} rows  {n_cols:2d} cols  '
            f'cond.fmt rules: {cf_rules}  freeze: {ws.freeze_panes}'
        )

    # Spot-check a CAGR cell fill
    console.rule('spot-check: CAGR fill on a few rows')
    ws = wb['CAGR_by_Institution']
    headers = [c.value for c in ws[1]]
    cagr_col = headers.index('CAGR') + 1
    flag_col = headers.index('Flag') + 1
    cagr_letter = get_column_letter(cagr_col)
    flag_letter = get_column_letter(flag_col)
    for r in range(2, min(ws.max_row, 8) + 1):
        cagr_cell = ws[f'{cagr_letter}{r}']
        flag_cell = ws[f'{flag_letter}{r}']
        fill = cagr_cell.fill.fgColor.rgb if cagr_cell.fill and cagr_cell.fill.fgColor else 'none'
        v = cagr_cell.value
        v_str = 'blank' if v is None else f'{v:.4f}'
        console.print(f'  row {r}: CAGR={v_str:8s}  flag={flag_cell.value or "":15s}  fill={fill}')
